#import socket
import telnetlib
import __main__
import openpyxl as px


def login ( hostname , address ):
    c = telnetlib.Telnet(address)
    c.read_until(b"Password:")
    c.write(b"cisco\n")
    c.read_until(bytes(hostname) + b">")
    c.write(b"enable\n")
    c.read_until(b"Password:")
    c.write(b"cisco\n")
    __main__.c = c


def length0 (hostname):
    c = __main__.c
    c.read_until(bytes(hostname) + b"#")
    c.write(b"ter len 0\n")    


def execute_command(hostname,command):
    c = __main__.c
    c.read_until(bytes(hostname) + b"#")
    c.write(bytes(command))
    result = c.read_until(bytes(hostname) + b"#")
    logname = "./" + hostname + ".cfg"
    wb = open(logname ,'wb')
    wb.write(result.decode('ascii'))
    print(result.decode('ascii'))

def close_connection():
    c = __main__.c
    c.write(b"q\n")

def get_excel_hostaddr():
    wb = px.load_workbook('testbook.xlsx')
    ws = wb.active
    i = 1
    dct = {}
    while ws['A' + str(i)].value != None :
        dct[ws['A' + str(i)].value] = ws['B' + str(i)].value
        i+=1
    __main__.dct = dct















get_excel_hostaddr()
key_lst = list(dct.keys())
print (key_lst[0])
print (dct[key_lst[0]])

#### rupe
login( key_lst[0], dct[key_lst[0]] )
length0(key_lst[0])
execute_command( (key_lst[0]),"show run")
close_connection()
#####
exit()





