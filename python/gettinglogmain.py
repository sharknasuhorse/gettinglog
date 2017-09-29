import pexpect
import sys
import time
import __main__
import openpyxl as px

#import os.path
#import xlrd
#ipaddr = 172.17.50.254
#hostname =
#logname

expect_list = [u"#",
                u">",
                u"\nlogin: ",
                u"Username: ",
                u"Password: ",
                u"Connection closed by foreign host.",
                u"Login incorrect"]


def login ():
    c = pexpect.spawn("telnet 172.17.50.254")
    c.expect("Password:")
    c.sendline("cisco")
    c.expect("2911_RT-A>")
    c.sendline("enable")
    c.expect("Password:")
    c.sendline("enable")
    c.sendline("cisco")
    __main__.c = c

def length0 ():
    c = __main__.c
    c.expect("2911_RT-A#")
    c.sendline("ter len 0")

def excute_command(command):
    c = __main__.c
    c.expect("2911_RT-A#")
    c.sendline(command)
    c.expect("2911_RT-A#")
    time.sleep(3)
    print(c.before)
    logname = "./" + "2911_RT-A" + ".cfg"
    wb = open(logname ,'wb')
    wb.write(c.before)

def close_settion():
    c = __main__.c
    c.sendline("q")
    c.close()

def get_excel_hostaddr():
    wb = px.load_workbook('testbook.xlsx')
    ws = wb.active
    i = 1
    dct = {}
    while ws['A' + str(i)].value != None :
        dct[ws['A' + str(i)].value] = ws['B' + str(i)].value
        i+=1

get_excel_hostaddr()

exit()

login()
length0()
excute_command("show run")
close_settion()

#logname = "./log" + "test" + ".log"
#wb = open(logname ,'wb')
#c.logfile_read = wb





